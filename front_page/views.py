# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import openpyxl
from django.shortcuts import render, render_to_response
from django.views.decorators.csrf import ensure_csrf_cookie
@ensure_csrf_cookie

def index(request):
    return render(request, "index.html", {})

def ml_index(request):
    import pandas as pd  
    import numpy as np  
    import pickle
    import matplotlib.pyplot as plt 
    # excel_file = request.FILES["excel_file"]
    df = pd.read_csv(request.FILES['excel_file'])
    # print(df.iloc[:,1])
    print(len(df.columns))
    for i in range(1,len(df.columns)-1):
        l = list(df.iloc[:,i])
        l1 = list(df.iloc[:,i+1])
        for j in range(0,len(l)):
            if(l1[j] == 0 or l[j] == 0):
                l[j]=0
            else:
                l[j] = ((l[j]-l1[j])/l1[j])*100
    #     print(l)
        df.iloc[:,i] = l
    df
    print(df)
    # df = df.transpose()
    df = df.iloc[:, :-1]
    df = df.transpose()
    df = df.drop([1,3],axis=1)
    df = df.drop('Unnamed: 0')
    # Exploratory Data Analysis
    print(df.shape)  
    print("------------")
    print(df.head()) 
    X_test = df
    loaded_model = pickle.load(open('/home/rahil/Desktop/FInal Project/finalized_model_7class.sav', 'rb'))
    y_pred=loaded_model.predict(X_test)

    df[0]
    years = list(df.index)
    company_performance = list(y_pred)
    company_performance.reverse()
    company_performance
    print(years)
    years.reverse()
    print(years)
    company_data = {}
    for i in range(0,len(years)):
        company_data[years[i]]=int(company_performance[i])
    years_number = []
    for i in range(1,len(years)+1):
        years_number.append(i)
    print(years_number)
    from numpy  import array
    years_number = array(years_number)
    company_performance = array(company_performance)
    n = np.size(years_number) 
    # mean of x and y vector 
    m_x, m_y = np.mean(years_number), np.mean(company_performance) 
    # calculating cross-deviation and deviation about x 
    SS_xy = np.sum(company_performance*years_number) - n*m_y*m_x 
    SS_xx = np.sum(years_number*years_number) - n*m_x*m_x 
    # calculating regression coefficients 
    m = SS_xy / SS_xx 
    c = m_y - m*m_x 
    np.mean(company_performance)
    import math
    change = math.sin(math.atan(m))
    value = c * change * np.size(years_number) * -1
    perc = ((7-value)/7) +  1
    final_grade = np.mean(company_performance)*perc
    # company_data['final_grade']=final_grade
    percentage = int(((7-final_grade)*100)/7)
    grade = {"final_grade": percentage}
    # company_data
    # print("Grade : ",grade)
    return render(request, "company_grade.html",{"company_data" : company_data, "grade" : grade} )

def excel_index(request):
    if "GET" == request.method:
        return render(request, "excel.html", {})
    else:
        excel_file = request.FILES["excel_file"]
        df1 = openpyxl.load_workbook(excel_file)
        df = df1["Sheet1"]
        current_ratio = []
        quick_ratio = []
        debt_equity = []
        current_ratio.append('current_ratio')
        quick_ratio.append('quick_ratio')
        debt_equity.append('debt_equity')

        shareholders = []
        Total_Assets = []
        Total_Liabilities = []
        cash = []
        investments = []
        current_assets = []
        receivables = []
        years = []
        # print(df)
        max_column = df.max_column
        for i in range(1, max_column + 1):
            shareholders.append(df.cell(row=7, column=i))
            Total_Assets.append(df.cell(row=36, column=i))
            years.append(df.cell(row=1, column=i).value)
            Total_Liabilities.append(df.cell(16, i))
            cash.append(df.cell(33, i))
            investments.append(df.cell(30, i))
            current_assets.append(df.cell(35, i))
            receivables.append(df.cell(32, i))

        shareholders = list(shareholders)
        Total_Assets = list(Total_Assets)
        # Total_Liabilities = list(df.iloc[14])
        # cash = list(df.iloc[31])
        # investments = list(df.iloc[28])
        # current_assets = list(df.iloc[33])
        # receivables = list(df.iloc[30])
        years.pop(0)
        cash.pop(0)
        investments.pop(0)
        current_assets.pop(0)
        receivables.pop(0)
        Total_Assets.pop(0)
        Total_Liabilities.pop(0)
        shareholders.pop(0)
        # year = []
        x_quick = cash + investments + current_assets + receivables
        print(x_quick)
        for i in range(0, len(Total_Liabilities)):
            # year.append(years[i].value)
            current_ratio.append(float(Total_Assets[i].value) / float(Total_Liabilities[i].value))
            x_quick[i] = float(cash[i].value) + float(investments[i].value) + float(current_assets[i].value) + float(
                receivables[i].value)
            quick_ratio.append(float(x_quick[i]) / float(Total_Liabilities[i].value))
            debt_equity.append(float(Total_Liabilities[i].value) / float(shareholders[i].value))
        print(current_ratio)
        print(quick_ratio)

        excel_data = list()

        excel_data.append(current_ratio)
        excel_data.append(quick_ratio)
        excel_data.append(debt_equity)
        print(len(excel_data))

        df2 = df1.create_sheet('Financial_Ratios')
        for row in excel_data:
            df2.append(row)
        df1.save(excel_file)

        print(df2.max_column)
        l3 = []
        current_ratio_perc = []
        l3.append(['current_ratio', 'quick_ratio', 'debt_equity'])
        excel_data = list(excel_data)
        for i in range(1, df2.max_column - 1):
            l = []
            l1 = []
            [l.append(excel_data[p][i]) for p in range(0, 3)]
            [l1.append(excel_data[k][i + 1]) for k in range(0, 3)]
            for j in range(0, len(l)):
                if (l1[j] == 0 or l[j] == 0):
                    l[j] = 0
                else:
                    l[j] = ((l[j] - l1[j]) / l1[j]) * 100
            l3.append(l)
            current_ratio_perc = [[l3[j][i] for j in range(len(l3))] for i in range(len(l3[0]))]




        # made a dictionary and added key name with a list value
        ratio_data = {"current_ratio": [], "quick_ratio": [], "debt_equity": [],"current_ratio_perc" : [],"quick_ratio_perc": [],"debt_equity_perc":[],"years":[]}
        i=1
        name = 0

        # we have a number of lists in excel_data and this function is
        # converting it into lists of each key based on list number
        for row in excel_data:
            i=0
            name=name+1
            for cell in row:
               if(i==0):
                   i=i+1
               elif(name==1):
                   ratio_data['current_ratio'].append(cell)
               elif(name==2):
                   ratio_data['quick_ratio'].append(cell)
               elif (name==3):
                   ratio_data['debt_equity'].append(cell)
        # for row in current_ratio_perc:
        #     ratio_data['current_ratio_perc'].append(row)
        name = 0
        for row in current_ratio_perc:
            i=0
            name=name+1
            for cell in row:
               if(i==0):
                   i=i+1
               elif(name==1):
                   ratio_data['current_ratio_perc'].append(cell)
               elif(name==2):
                   ratio_data['quick_ratio_perc'].append(cell)
               elif (name==3):
                   ratio_data['debt_equity_perc'].append(cell)
            # ratio_data['current_ratio_perc'].append(row)
        ratio_data['years'].append(years)
        print(years)
        # passing a dictionary to the template
        return render(request, 'excel.html', ratio_data)



